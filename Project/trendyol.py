import requests
from bs4 import BeautifulSoup
import re
from main import Product
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from selenium.common.exceptions import NoSuchElementException
import time


product = Product()


def edit_search(search):
    url = "https://www.trendyol.com/sr?q=" 
    s_words = search.split(" ")
    len_search = len(search.split(" "))
    
    for i  in range(len_search-1):
        url = url + s_words[i]
        url += "%20"
    url += s_words[len_search-1]
    url += "&qt="
    for i  in range(len_search-1):
        url = url + s_words[i]
        url += "%20"
    url += s_words[len_search-1]
    url += "&st="
    for i  in range(len_search-1):
        url = url + s_words[i]
        url += "%20"
    url += s_words[len_search-1]
    url += "&os=1"

    edit_html(url)


def edit_html(url):
    sayfa = requests.get(url)
    html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
    isim = html_sayfa.find_all("div", class_="p-card-chldrn-cntnr card-border")
    isim = isim[:10]

    with open("links.txt", "w") as file:
        for i in range(len(isim)):
            x = f"{isim[i]}\n"
            file.write(x)

    with open("links.txt", "r") as file:
        lines = file.readlines()
    with open("links.txt", "w") as file:
        for line in lines:
            line = re.sub(r'(<div class="p-card-chldrn-cntnr card-border"><a href=")', "", line)
            line = re.sub(r'("><div class=").*', "", line)
            line = re.sub(r'amp;', "", line)
            line = "https://www.trendyol.com" + line
            file.writelines(line)
    


def take_product_info():
    with open("links.txt", "r") as file:
        lines = file.readlines()
    create_excel()
    x = 2
    for line in lines:
        time.sleep(2)
        product = Product(take_kategori(line), take_marka(line), take_ilanismi(line), take_fiyat(line), take_seller(line), take_variants(line), take_ratings(line), take_reviews(line), take_answers(line), take_genel(line), take_other_sellers(line))
        edit_excel(product, x, line)
        x += 1


def take_kategori(link):
    sayfa = requests.get(link)
    html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
    isim = html_sayfa.find("div", class_="product-detail-breadcrumb full-width").getText()
    isim = re.sub(r'([A-Z])', r' \1', isim)
    return str(isim)

def take_marka(link):
    sayfa = requests.get(link)
    html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
    try:
        isim = html_sayfa.find("a", class_="product-brand-name-with-link").getText()
        return isim
    except AttributeError:
        isim = html_sayfa.find("span", class_="product-brand-name-without-link").getText()
        return isim

def take_ilanismi(link):
    sayfa = requests.get(link)
    html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
    isim = html_sayfa.find("h1", class_="pr-new-br").getText()
    return str(isim)

def take_fiyat(link):
    try:
        sayfa = requests.get(link)
        html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
        isim = html_sayfa.find("div", class_="pr-bx-nm with-org-prc").getText()
        return str(isim)
    except AttributeError:
        sayfa = requests.get(link)
        html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
        isim = html_sayfa.find("div", class_="featured-prices").getText()
        return str(isim)
    
def take_seller(link):
    sayfa = requests.get(link)
    html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
    isim = html_sayfa.find("div", class_="flex-container").getText()
    seller = re.search(r"Bu ürün (.+?) tarafından", isim).group(1)
    return str(seller)

def take_variants(link):
    pass

def take_ratings(link):
    x = ""
    try:
        service = Service("./chromediver.exe")
        driver = webdriver.Chrome(service=service)
        driver.minimize_window()
        driver.get(link)
        result = driver.find_element(By.CLASS_NAME, "pr-in-rnr")
        res = result.text.splitlines()
        x += f"{res[0]} Satıcı Puanı "
        for i in range(len(res)-1):
            x += f"\n{res[i+1]} "
        return x
    except NoSuchElementException:
        return "There's no ratings"

def take_reviews(link):
    service = Service("./chromediver.exe")
    driver = webdriver.Chrome(service=service)
    link = re.sub(r"\?boutiqueId=", "/yorumlar\?boutiqueId=", link)
    driver.minimize_window()
    driver.get(link)

    result = driver.find_elements(By.CLASS_NAME, "comment")

    son = ""
    x = 1
    for text in result:
        t = text.text
        t = t.splitlines()
        son += (str(x) + ". -> " + t[2] + "-" + t[3] + "\n")
        x += 1
    if son != "":
        return son
    else:
        return "No review for this product"

def take_answers(link):
    service = Service("./chromediver.exe")
    driver = webdriver.Chrome(service=service)
    link = re.sub(r"\?boutiqueId=.*&", "/saticiya-sor?", link)
    driver.minimize_window()
    driver.get(link)
    if "Bulunamadı" in driver.find_element(By.CLASS_NAME, "pr-qna-v2").text:
        return ("There's no answers for this seller")
    else:
        result = driver.find_elements(By.CLASS_NAME, "item-content")
        answers = []
        for i in range(len(result)):
            text = result[i].text
            x = text.splitlines()
            answers.append(x[0])

        result2 = driver.find_elements(By.CLASS_NAME, "answer")
        replies = []
        for i in range(len(result)):
            text = result2[i].text
            x = text.splitlines()
            cevap = x[0]+x[1]
            cevap = re.sub(r".*içinde cevaplandı.", "", cevap)
            replies.append(cevap)
        son = ""
        for i in range(len(answers)):
            son += (f"Soru{i+1}: {answers[i]} - Cevap{i+1}: {replies[i]}\n")
        return (son)

def take_genel(link):
    sayfa = requests.get(link)
    html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
    isim = html_sayfa.find("div", class_="flex-container").getText()
    return str(isim)

def take_other_sellers(link):
    sayfa = requests.get(link)
    html_sayfa = BeautifulSoup(sayfa.content, "html.parser")
    isim = html_sayfa.find_all("div", class_="pr-mc-w gnr-cnt-br")
    text = "(Satıcı Adı/Satıcı Puanı/Kargo/Fatura/Fiyat)\n\n"

    x = BeautifulSoup(str(isim), "html.parser")
    linkler = []
    for i in x.find_all("a"):
        linkler.append(i['href'])

    for i in range(len(isim)):
        o = BeautifulSoup(isim[i].text, "html.parser")
        o = re.sub(r"Ürüne Git", "", str(o))
        text += f"{i+2}. satıcı: " + str(o) + " https://www.trendyol.com/" + linkler[i] + "\n"
    return text


##Main function
def main(search):
    edit_search(search)
    take_product_info()
    

def create_excel():
    file = Workbook()
    sheet = file.active
    sheet['A1'] = 'Categori'
    sheet['B1'] = 'Brand'
    sheet['C1'] = 'Ad Name'
    sheet['D1'] = 'Price'
    sheet['E1'] = 'Seller'
    sheet['F1'] = 'Variants'
    sheet['G1'] = 'Ratings'
    sheet['H1'] = 'Reviews'
    sheet['I1'] = 'ANSWER-REPLY'
    sheet['J1'] = 'General'
    sheet['K1'] = 'E-Commerce Site'
    sheet['L1'] = 'Link'
    sheet['M1'] = 'Other Sellers'

    bold_font = Font(bold=True)
    fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
    cell_range = 'A1:M1'
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.font = bold_font
    file.save("products.xlsx")


def edit_excel(product, x, link):
    file = load_workbook("products.xlsx")
    sheet = file.active

    sheet[f"A{x}"].value = f'{product.kategori}'
    sheet[f"B{x}"].value = f'{product.marka}'
    sheet[f"C{x}"].value = f'{product.ilanismi}'
    sheet[f"D{x}"].value = f'{product.fiyat}'
    sheet[f"E{x}"].value = f'{product.seller}'
    sheet[f"F{x}"].value = f'{product.variants}'
    sheet[f"G{x}"].value = f'{product.ratings}'
    sheet[f"H{x}"].value = f'{product.reviews}'
    sheet[f"I{x}"].value = f'{product.answers}'
    sheet[f"J{x}"].value = f'{product.genel}'
    sheet[f"K{x}"].value = 'TRENDYOL'
    sheet[f"L{x}"].value = f'{link}'
    sheet[f"M{x}"].value = f'{product.other_sellers}'

    file.save("products.xlsx")
